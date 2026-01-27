"""
Create an example Excel spreadsheet demonstrating the Fackler Distributions add-in.

This script generates an Excel file with:
- Worksheets for each distribution
- Parameter inputs
- PDF and CDF formulas using the add-in functions
- Charts for each distribution

Usage:
    python create_excel_demo.py

Output:
    dist/FacklerDistributions_Demo.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import numpy as np


# Output directory
OUTPUT_DIR = "dist"


def setup_output_dir():
    """Create output directory if it doesn't exist."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)


# Styles
HEADER_FONT = Font(bold=True, size=14, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
PARAM_FILL = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
FORMULA_FONT = Font(name="Consolas", size=10)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_distribution_sheet(wb, name, title, description, params, pdf_formula, cdf_formula,
                               sf_formula=None, x_range=(0.1, 20), num_points=100,
                               extra_info=None, derived_formulas=None):
    """
    Create a worksheet for a distribution.

    Args:
        wb: Workbook
        name: Sheet name
        title: Distribution title
        description: Description text
        params: List of (param_name, param_label, default_value) tuples
        pdf_formula: Formula template for PDF (use {x} and param names)
        cdf_formula: Formula template for CDF
        sf_formula: Formula template for SF (optional)
        x_range: (x_min, x_max) for data
        num_points: Number of data points
        extra_info: Additional information text
        derived_formulas: List of (label, formula) for derived parameters
    """
    ws = wb.create_sheet(title=name)

    # Title
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    # Description
    ws['A3'] = "Description:"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = description
    ws.merge_cells('A4:C6')
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')

    # Parameters section
    ws['A8'] = "Parameters"
    ws['A8'].font = HEADER_FONT
    ws['A8'].fill = HEADER_FILL
    ws.merge_cells('A8:B8')

    param_cells = {}
    row = 9
    for param_name, param_label, default_value in params:
        ws[f'A{row}'] = param_label
        ws[f'A{row}'].fill = PARAM_FILL
        ws[f'B{row}'] = default_value
        ws[f'B{row}'].fill = PARAM_FILL
        ws[f'B{row}'].border = THIN_BORDER
        param_cells[param_name] = f'$B${row}'
        row += 1

    # Derived parameters (if any)
    if derived_formulas:
        row += 1
        ws[f'A{row}'] = "Derived Parameters"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for label, formula in derived_formulas:
            ws[f'A{row}'] = label
            # Replace param names with cell references
            formula_with_refs = formula
            for pname, pcell in param_cells.items():
                formula_with_refs = formula_with_refs.replace(f'{{{pname}}}', pcell)
            ws[f'B{row}'] = formula_with_refs
            ws[f'B{row}'].font = FORMULA_FONT
            row += 1

    # Extra info
    if extra_info:
        row += 1
        ws[f'A{row}'] = extra_info
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:C{row}')

    # Data section
    data_start_row = 9
    data_col_x = 'D'
    data_col_pdf = 'E'
    data_col_cdf = 'F'
    data_col_sf = 'G'

    ws[f'{data_col_x}{data_start_row - 1}'] = "x"
    ws[f'{data_col_pdf}{data_start_row - 1}'] = "PDF"
    ws[f'{data_col_cdf}{data_start_row - 1}'] = "CDF"
    ws[f'{data_col_sf}{data_start_row - 1}'] = "SF"

    for col in [data_col_x, data_col_pdf, data_col_cdf, data_col_sf]:
        ws[f'{col}{data_start_row - 1}'].font = HEADER_FONT
        ws[f'{col}{data_start_row - 1}'].fill = HEADER_FILL

    # Generate x values
    x_min, x_max = x_range
    x_values = np.linspace(x_min, x_max, num_points)

    for i, x_val in enumerate(x_values):
        row = data_start_row + i
        x_cell = f'{data_col_x}{row}'

        ws[x_cell] = round(x_val, 6)

        # Build formulas with cell references
        pdf_f = pdf_formula.replace('{x}', x_cell)
        cdf_f = cdf_formula.replace('{x}', x_cell)

        for pname, pcell in param_cells.items():
            pdf_f = pdf_f.replace(f'{{{pname}}}', pcell)
            cdf_f = cdf_f.replace(f'{{{pname}}}', pcell)

        ws[f'{data_col_pdf}{row}'] = pdf_f
        ws[f'{data_col_cdf}{row}'] = cdf_f

        if sf_formula:
            sf_f = sf_formula.replace('{x}', x_cell)
            for pname, pcell in param_cells.items():
                sf_f = sf_f.replace(f'{{{pname}}}', pcell)
            ws[f'{data_col_sf}{row}'] = sf_f
        else:
            ws[f'{data_col_sf}{row}'] = f'=1-{data_col_cdf}{row}'

    data_end_row = data_start_row + num_points - 1

    # Create PDF chart
    pdf_chart = LineChart()
    pdf_chart.title = f"{name} - PDF"
    pdf_chart.style = 10
    pdf_chart.x_axis.title = "x"
    pdf_chart.y_axis.title = "f(x)"
    pdf_chart.width = 15
    pdf_chart.height = 10

    pdf_data = Reference(ws, min_col=5, min_row=data_start_row - 1,
                          max_col=5, max_row=data_end_row)
    x_data = Reference(ws, min_col=4, min_row=data_start_row,
                        max_col=4, max_row=data_end_row)
    pdf_chart.add_data(pdf_data, titles_from_data=True)
    pdf_chart.set_categories(x_data)

    ws.add_chart(pdf_chart, "I8")

    # Create CDF chart
    cdf_chart = LineChart()
    cdf_chart.title = f"{name} - CDF & SF"
    cdf_chart.style = 10
    cdf_chart.x_axis.title = "x"
    cdf_chart.y_axis.title = "F(x)"
    cdf_chart.width = 15
    cdf_chart.height = 10

    cdf_data = Reference(ws, min_col=6, min_row=data_start_row - 1,
                          max_col=7, max_row=data_end_row)
    cdf_chart.add_data(cdf_data, titles_from_data=True)
    cdf_chart.set_categories(x_data)

    ws.add_chart(cdf_chart, "I25")

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    return ws


def create_intro_sheet(wb):
    """Create introduction sheet."""
    ws = wb.active
    ws.title = "Introduction"

    ws['A1'] = "Fackler Distributions Excel Add-in"
    ws['A1'].font = Font(bold=True, size=20)
    ws.merge_cells('A1:G1')

    ws['A3'] = "Based on: Fackler, M. (2013) 'Reinventing Pareto: Fits for both small and large losses'"
    ws['A3'].font = Font(italic=True, size=12)
    ws.merge_cells('A3:G3')

    content = """
This Excel add-in provides functions for actuarial loss severity distributions
from the Pareto family and their extensions.

INSTALLATION:
1. Close Excel
2. Double-click the FacklerDistributions-AddIn64.xll file
3. When Excel opens, click "Enable" if prompted
4. The functions will be available with prefix FACKLER.*

AVAILABLE DISTRIBUTIONS:

Core Pareto Family:
  - FACKLER.PARETO.*    : Pareto Type I (Single-parameter)
  - FACKLER.GPD.*       : Generalized Pareto Distribution
  - FACKLER.LOMAX.*     : Lomax (Pareto Type II)
  - FACKLER.BURR.*      : Burr Type XII

Spliced Distributions:
  - FACKLER.LNPARETO.*  : Lognormal-Pareto (Czeledin distribution)
  - FACKLER.WEIPARETO.* : Weibull-Pareto
  - FACKLER.EXPPARETO.* : Exponential-Pareto
  - FACKLER.POWPARETO.* : Power-Pareto (Log-Laplace)

Mixed Distributions:
  - FACKLER.PARETO4.*   : 4-Parameter Pareto
  - FACKLER.PARETO5.*   : 5-Parameter Pareto

Each distribution has .PDF, .CDF, and .SF (survival function) variants.

REFERENCES:
  - Fackler (2013): reference/Fackler_Paper_Hague2013.pdf
  - Wikipedia: Pareto, GPD, Lomax, Burr distributions

See the other sheets in this workbook for examples of each distribution.
"""

    row = 5
    for line in content.strip().split('\n'):
        ws[f'A{row}'] = line
        row += 1

    ws.column_dimensions['A'].width = 100


def create_workbook():
    """Create the complete demo workbook."""
    wb = Workbook()

    # Introduction sheet
    create_intro_sheet(wb)

    # Pareto Type I
    create_distribution_sheet(
        wb, "Pareto",
        "Pareto Type I Distribution",
        "The classic reinsurance tail model. F̄(x) = (θ/x)^α for x ≥ θ. "
        "Has a 'memoryless' property where α remains invariant across different thresholds.",
        params=[
            ('alpha', 'α (alpha)', 2.0),
            ('theta', 'θ (theta)', 1.0),
        ],
        pdf_formula='=FACKLER.PARETO.PDF({x},{alpha},{theta})',
        cdf_formula='=FACKLER.PARETO.CDF({x},{alpha},{theta})',
        sf_formula='=FACKLER.PARETO.SF({x},{alpha},{theta})',
        x_range=(1.0, 10),
        extra_info="Note: x must be ≥ θ. For x < θ, PDF=0 and CDF=0."
    )

    # GPD
    create_distribution_sheet(
        wb, "GPD",
        "Generalized Pareto Distribution",
        "GPD generalizes Pareto with F̄(x|X>θ) = ((θ+λ)/(x+λ))^α. "
        "Both α and λ are invariant when modeling higher tails.",
        params=[
            ('alpha', 'α (alpha)', 2.0),
            ('lambda', 'λ (lambda)', 1.0),
            ('theta', 'θ (theta)', 1.0),
        ],
        pdf_formula='=FACKLER.GPD.PDF({x},{alpha},{lambda},{theta})',
        cdf_formula='=FACKLER.GPD.CDF({x},{alpha},{lambda},{theta})',
        sf_formula='=FACKLER.GPD.SF({x},{alpha},{lambda},{theta})',
        x_range=(1.0, 15),
        extra_info="λ=0 reduces to Pareto. λ>0 with θ=0 gives Lomax."
    )

    # Lomax
    create_distribution_sheet(
        wb, "Lomax",
        "Lomax Distribution (Pareto Type II)",
        "Lomax is a special case of GPD with θ=0. F̄(x) = (λ/(x+λ))^α. "
        "A ground-up distribution starting at 0.",
        params=[
            ('alpha', 'α (alpha)', 2.0),
            ('lambda', 'λ (lambda)', 2.0),
        ],
        pdf_formula='=FACKLER.LOMAX.PDF({x},{alpha},{lambda})',
        cdf_formula='=FACKLER.LOMAX.CDF({x},{alpha},{lambda})',
        sf_formula='=FACKLER.LOMAX.SF({x},{alpha},{lambda})',
        x_range=(0.01, 15)
    )

    # Burr
    create_distribution_sheet(
        wb, "Burr",
        "Burr Type XII Distribution",
        "Three-parameter generalization. F̄(x) = (1/(1+(x/λ)^τ))^α. "
        "τ>1 gives bell-shaped density, τ<1 gives decreasing density.",
        params=[
            ('alpha', 'α (alpha)', 2.0),
            ('lambda', 'λ (lambda)', 2.0),
            ('tau', 'τ (tau)', 2.0),
        ],
        pdf_formula='=FACKLER.BURR.PDF({x},{alpha},{lambda},{tau})',
        cdf_formula='=FACKLER.BURR.CDF({x},{alpha},{lambda},{tau})',
        sf_formula='=FACKLER.BURR.SF({x},{alpha},{lambda},{tau})',
        x_range=(0.01, 10),
        extra_info="Asymptotic Pareto exponent = α·τ"
    )

    # Lognormal-Pareto
    create_distribution_sheet(
        wb, "LN-Pareto",
        "Lognormal-Pareto (Czeledin) Distribution",
        "Spliced model: Lognormal body for x<θ, Pareto tail for x≥θ. "
        "PDF is continuous at θ. Popular in reinsurance.",
        params=[
            ('mu', 'μ (mu)', 1.0),
            ('sigma', 'σ (sigma)', 0.8),
            ('theta', 'θ (theta)', 10.0),
        ],
        pdf_formula='=FACKLER.LNPARETO.PDF({x},{mu},{sigma},{theta})',
        cdf_formula='=FACKLER.LNPARETO.CDF({x},{mu},{sigma},{theta})',
        sf_formula='=FACKLER.LNPARETO.SF({x},{mu},{sigma},{theta})',
        x_range=(0.1, 50),
        derived_formulas=[
            ('Derived α', '=FACKLER.LNPARETO.ALPHA({mu},{sigma},{theta})'),
        ],
        extra_info="α is derived from the C1 smoothness condition."
    )

    # Weibull-Pareto
    create_distribution_sheet(
        wb, "Wei-Pareto",
        "Weibull-Pareto Distribution",
        "Spliced model: Weibull body for x<θ, Pareto tail for x≥θ. "
        "k>1 gives bell-shaped body, k<1 gives decreasing body.",
        params=[
            ('k', 'k (shape)', 2.0),
            ('scale', 'scale', 5.0),
            ('theta', 'θ (theta)', 10.0),
            ('r', 'r (body weight)', 0.7),
        ],
        pdf_formula='=FACKLER.WEIPARETO.PDF({x},{k},{scale},{theta},{r})',
        cdf_formula='=FACKLER.WEIPARETO.CDF({x},{k},{scale},{theta},{r})',
        sf_formula='=FACKLER.WEIPARETO.SF({x},{k},{scale},{theta},{r})',
        x_range=(0.1, 40),
        derived_formulas=[
            ('Derived α', '=FACKLER.WEIPARETO.ALPHA({k},{scale},{theta},{r})'),
        ]
    )

    # Exponential-Pareto
    create_distribution_sheet(
        wb, "Exp-Pareto",
        "Exponential-Pareto Distribution",
        "Spliced model with only 2 parameters. Exponential body, Pareto tail. "
        "Local α increases linearly from 0 to α across the body.",
        params=[
            ('rate', 'rate (λ)', 0.15),
            ('theta', 'θ (theta)', 10.0),
        ],
        pdf_formula='=FACKLER.EXPPARETO.PDF({x},{rate},{theta})',
        cdf_formula='=FACKLER.EXPPARETO.CDF({x},{rate},{theta})',
        sf_formula='=FACKLER.EXPPARETO.SF({x},{rate},{theta})',
        x_range=(0.01, 40),
        derived_formulas=[
            ('Derived α', '=FACKLER.EXPPARETO.ALPHA({rate},{theta})'),
            ('Derived r', '=FACKLER.EXPPARETO.R({rate},{theta})'),
        ],
        extra_info="α = θ·rate, r = 1-exp(-rate·θ)"
    )

    # Power-Pareto
    create_distribution_sheet(
        wb, "Pow-Pareto",
        "Power-Pareto (Log-Laplace) Distribution",
        "Spliced model: Power function body (x/θ)^β, Pareto tail. "
        "Also known as double Pareto or Log-Laplace distribution.",
        params=[
            ('alpha', 'α (alpha)', 2.0),
            ('beta', 'β (beta)', 1.5),
            ('theta', 'θ (theta)', 5.0),
        ],
        pdf_formula='=FACKLER.POWPARETO.PDF({x},{alpha},{beta},{theta})',
        cdf_formula='=FACKLER.POWPARETO.CDF({x},{alpha},{beta},{theta})',
        sf_formula='=FACKLER.POWPARETO.SF({x},{alpha},{beta},{theta})',
        x_range=(0.01, 20),
        derived_formulas=[
            ('Derived r', '=FACKLER.POWPARETO.R({alpha},{beta})'),
        ],
        extra_info="r = α/(α+β) from C1 smoothness condition"
    )

    # 4-Parameter Pareto
    create_distribution_sheet(
        wb, "Pareto-4P",
        "4-Parameter Pareto Distribution",
        "Mixture of two Lomax distributions with constraint α₁ = α₂ + 2. "
        "Used by ISO for general liability business.",
        params=[
            ('alpha2', 'α₂', 1.5),
            ('lambda1', 'λ₁', 2.0),
            ('lambda2', 'λ₂', 10.0),
            ('r', 'r (weight)', 0.6),
        ],
        pdf_formula='=FACKLER.PARETO4.PDF({x},{alpha2},{lambda1},{lambda2},{r})',
        cdf_formula='=FACKLER.PARETO4.CDF({x},{alpha2},{lambda1},{lambda2},{r})',
        sf_formula='=FACKLER.PARETO4.SF({x},{alpha2},{lambda1},{lambda2},{r})',
        x_range=(0.01, 50),
        extra_info="α₁ = α₂ + 2 (constraint)"
    )

    # 5-Parameter Pareto
    create_distribution_sheet(
        wb, "Pareto-5P",
        "5-Parameter Pareto Distribution",
        "Unconstrained mixture of two Lomax distributions. "
        "First component dominates small losses, second dominates tail.",
        params=[
            ('alpha1', 'α₁', 3.0),
            ('alpha2', 'α₂', 1.5),
            ('lambda1', 'λ₁', 2.0),
            ('lambda2', 'λ₂', 15.0),
            ('r', 'r (weight)', 0.7),
        ],
        pdf_formula='=FACKLER.PARETO5.PDF({x},{alpha1},{alpha2},{lambda1},{lambda2},{r})',
        cdf_formula='=FACKLER.PARETO5.CDF({x},{alpha1},{alpha2},{lambda1},{lambda2},{r})',
        sf_formula='=FACKLER.PARETO5.SF({x},{alpha1},{alpha2},{lambda1},{lambda2},{r})',
        x_range=(0.01, 80)
    )

    return wb


def main():
    """Generate the demo workbook."""
    print("Creating Fackler Distributions Demo Workbook...")
    setup_output_dir()

    wb = create_workbook()

    output_path = os.path.join(OUTPUT_DIR, "FacklerDistributions_Demo.xlsx")
    wb.save(output_path)

    print(f"Saved: {output_path}")
    print("\nNote: The formulas in this workbook use FACKLER.* functions.")
    print("You must install the Excel add-in for the formulas to work.")
    print("Without the add-in, cells will show #NAME? errors.")


if __name__ == "__main__":
    main()
